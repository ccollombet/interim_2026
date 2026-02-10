import streamlit as st
import pandas as pd
import os
import tempfile
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy
import re
from datetime import datetime
import unicodedata
import logging

# ======================================================
# CONFIG STREAMLIT (UNE SEULE FOIS)
# ======================================================
st.set_page_config(
    page_title="Générateur de planning – Pipeline complet",
    layout="centered"
)
st.title("🗓️ Générateur de planning – Pipeline complet")

# ======================================================
# HELPERS COMMUNS
# ======================================================
#MOTIFS_PREDEFINIS = [
#    "Absence maladie",
#    "Accroissement temporaire d'activité",
#
#    "Formation",
#    "Arrêt Maladie",
#    "Accident de travail",
#    "Absence injustifiée",
#    "Congés",
#    "Congé d'ancienneté",
#    "Congés Payés",
#    "Congés sans solde",
#    "Congés spécifiques/trimestriels",
#   "Surcroît temporaire d'activité CNR"
#]

MOTIFS_PREDEFINIS = [
    "Accident de travail",
    "Arrêt Maladie",
    "Congé de Maternité",
    "Congé parental d'éducation",
    "Congés Payés",
    "Formation",
    "Mi-temps Thérapeutique",
    "Récupération",
    "Surcroît temporaire d'activité CNR ou",
    "Surcroit temporaire d’activité",
    "Absence injustifiée",
    "Congé d'ancienneté",
    "Congé de Paternité",
    "Congé de présence parentale",
    "Congé Individuel de Formation",
    "Congé sabbatique",
    "Congés Évènements Familiaux",
    "Congés sans solde",
    "Congés spécifiques/trimestriels",
    "Dans l'attente de la nomination du titulaire",
    "Détachement du titulaire sur une tâche exceptionnelle",
    "Mise à pied conservatoire",
    "Mise à pied disciplinaire",
    "Réduction temps travail femme enceinte"
]

def save_uploaded_file(uploaded_file, suffix):
    temp_dir = tempfile.mkdtemp()
    file_path = os.path.join(
        temp_dir,
        f"{Path(uploaded_file.name).stem}_{suffix}{Path(uploaded_file.name).suffix}"
    )
    with open(file_path, "wb") as f:
        f.write(uploaded_file.getbuffer())
    return file_path


# =========================
#   HELPERS GÉNÉRIQUES
# =========================
def noacc_lower(s: str) -> str:
    """Minuscule sans accents + espaces normalisés."""
    if s is None:
        return ""
    s = str(s)
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    s = s.replace("\u00a0", " ").replace("\ufeff", "")
    s = re.sub(r"[ \t]+", " ", s).strip().lower()
    return s

def norm_group(s: str) -> str:
    """Normalise un libellé de groupe (REMPLACANT 1 G1, etc.)."""
    s = noacc_lower(s).replace("\n", " ")
    s = re.sub(r"\s+", " ", s).strip()
    # homogénéise 'remplaçant'/'remplacant'
    s = s.replace("remplaçant", "remplacant")
    return s

def strip_placeholders(s: str) -> str:
    """Supprime 'Nom'/'Prénom' en tête de la chaîne."""
    if not isinstance(s, str):
        return ""
    s = s.strip()
    s = re.sub(r"^\s*(nom|pr[ée]nom)\s*[/:\-]?\s*", "", s, flags=re.IGNORECASE)
    while re.match(r"^(nom|pr[ée]nom)\b", s, flags=re.IGNORECASE):
        s = re.sub(r"^\s*(nom|pr[ée]nom)\s*[/:\-]?\s*", "", s, flags=re.IGNORECASE)
    return s.strip()

def save_uploaded_file(uploaded_file, suffix):
    temp_dir = tempfile.mkdtemp()
    file_path = os.path.join(
        temp_dir,
        f"{Path(uploaded_file.name).stem}_{suffix}{Path(uploaded_file.name).suffix}"
    )
    with open(file_path, "wb") as f:
        f.write(uploaded_file.getbuffer())
    return file_path

# =========================
#   DÉTECTION DES JOURS
# =========================
_MOIS_MAP = {
    "jan": "01", "fev": "02", "fév": "02", "mar": "03", "mars": "03",
    "avr": "04", "mai": "05", "juin": "06", "jun": "06",
    "jui": "07", "juil": "07",
    "aou": "08", "aout": "08", "août": "08",
    "sep": "09", "sept": "09",
    "oct": "10", "nov": "11",
    "dec": "12", "déc": "12"
}

def _norm_text(s: str) -> str:
    s = (s or "").strip().replace("\n", " ").replace(".", " ")
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).lower()

def parse_header_to_date(header_val, year: str) -> str | None:
    """Convertit une entête 'L02 Mars' -> '02/mm/YYYY' en utilisant year déduite."""
    if not isinstance(header_val, str) or not header_val.strip():
        return None
    s = _norm_text(header_val)
    m = re.search(r"(\d{1,2})\s*([a-z]{3,5})", s)
    if not m:
        return None
    j = int(m.group(1))
    mois_tok = m.group(2)[:4]
    if mois_tok.startswith("jui"):   # juillet
        mois_tok = "jui"
    if mois_tok.startswith("aou"):   # août
        mois_tok = "aou"
    if mois_tok.startswith("dec"):   # décembre
        mois_tok = "dec"
    mois = _MOIS_MAP.get(mois_tok)
    if not mois:
        return None
    return f"{j:02d}/{mois}/{year}"

def detect_day_columns_p1(ws, start_col=5):
    """Colonnes jours (ligne 1), à partir de E."""
    day_cols = []
    for col in range(start_col, ws.max_column + 1):
        raw = ws.cell(row=1, column=col).value
        if raw in (None, ""):
            break
        # year sera injectée plus tard via closure
        day_cols.append(col)
    # On ne filtre pas ici : on laisse la conversion faire foi plus tard
    return day_cols

def guess_year_from_column_a(ws, default_year="2026") -> str:
    """Déduit l'année depuis col.A (lignes 'dd/mm/yyyy : Nom ...')."""
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if isinstance(val, str):
            m = re.search(r"\b\d{2}/\d{2}/(\d{4})\b", val)
            if m:
                return m.group(1)
    return default_year

# =========================
#   NÉTOYAGES FEUILLE
# =========================
def is_placeholder_cell(val) -> bool:
    if not isinstance(val, str):
        return False
    t = noacc_lower(val)
    t = re.sub(r"[/: \t]+", "", t)
    return t in {"nom", "prenom"}

def nettoyer_nom_ligne4(ws, col_debut=5):
    """Supprime 'Nom/' résiduel en ligne 4."""
    for col in range(col_debut, ws.max_column + 1):
        v = ws.cell(row=4, column=col).value
        if isinstance(v, str) and noacc_lower(v) in {"nom", "nom/"}:
            ws.cell(row=4, column=col).value = None

def nettoyer_prenom_dans_ligne_nom(ws, start_col=5):
    """Si D='Nom', supprime 'Prénom' en E.. jours."""
    # colonnes jours fallback = toutes à partir de E
    day_cols = list(range(start_col, ws.max_column + 1))
    for r in range(1, ws.max_row + 1):
        if noacc_lower(ws.cell(row=r, column=4).value) == "nom":
            for c in day_cols:
                if is_placeholder_cell(ws.cell(row=r, column=c).value):
                    ws.cell(row=r, column=c).value = None

# =========================
#   EXTRACTION REMPLA (COL. A)
# =========================
DATE_LINE_RE = re.compile(r"^\s*(\d{1,2})/(\d{1,2})/(\d{4})\s*[:：]\s*(.+?)\s*$")

def extract_remplacants_from_colA(xlsx_path: str) -> pd.DataFrame:
    """
    Extrait les lignes 'dd/mm/yyyy : NOM PRENOM' associées
    aux blocs dont la CATEGORIE (colonne C) == 2.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    rows = []
    current_group_raw = None
    current_is_rempla = False

    for r in range(1, ws.max_row + 1):
        val_a = ws.cell(row=r, column=1).value
        val_c = ws.cell(row=r, column=3).value  # colonne C = CATEGORIE

        if val_a is None:
            continue

        raw_a = str(val_a).strip()

        # --------------------------------------------------
        # 1️⃣ Ligne "identité / groupe"
        # --------------------------------------------------
        # On considère que c’est une ligne identité si la colonne C est renseignée
        if val_c is not None:
            try:
                categorie = int(val_c)
            except Exception:
                categorie = None

            # 👉 NOUVELLE LOGIQUE
            current_is_rempla = (categorie == 2)

            if current_is_rempla:
                current_group_raw = raw_a.replace("\n", " ").strip()
            else:
                current_group_raw = None

            continue

        # --------------------------------------------------
        # 2️⃣ Ligne "date : personne"
        # --------------------------------------------------
        if not current_is_rempla or not current_group_raw:
            continue

        m = DATE_LINE_RE.match(raw_a)
        if not m:
            continue

        j, mth, y = m.group(1), m.group(2), m.group(3)
        person_raw = strip_placeholders(m.group(4).strip())

        tokens = [t for t in person_raw.split() if t]
        if tokens:
            nom = tokens[0].strip(",;")
            prenom = " ".join(tokens[1:]).strip(",;")
        else:
            nom, prenom = "", ""

        rows.append({
            "date": f"{int(j):02d}/{int(mth):02d}/{y}",
            "groupe": current_group_raw,
            "nom": nom,
            "prenom": prenom
        })

    return pd.DataFrame(rows, columns=["date", "groupe", "nom", "prenom"])

# ------------------------------------------------------------------
# Mapping codes structure → Nom structure
# - Supporte codes longs (675020902, 675010101…)
# - Supporte fallback sur les 3 derniers chiffres
# ------------------------------------------------------------------
def get_structure_mapping() -> dict[str, str]:
    data = [
        ("6750404", "EA ADAPAYSAGE BOURG"),
        ("6750405", "EA ADAPAYSAGE HAUT BUGEY"),
        ("6750309", "ESAT BELLEGARDE INDUSTRIE"),
        ("6750313", "ESAT CENTRE DE VIE RURALE"),
        ("6750307", "ESAT LA LECHERE"),
        ("6750305", "ESAT LE PENNESSUY"),
        ("6750311", "ESAT LES ATELIERS DE NIERME"),
        ("6750303", "ESAT LES BROSSES"),
        ("6750301", "ESAT LES DOMBES"),
        ("6750315", "ESAT LES TEPPES"),
        ("6750503", "FAM PRE LA TOUR"),
        ("6750504", "FAM SOUS LA ROCHE"),
        ("6750215", "FOYER BELLEVUE"),
        ("6750212", "FOYER DE TREFFORT"),
        ("6750213", "FOYER COURTES VERNOUX"),
        ("6750203", "FOYER CROIX BLANCHE"),
        ("6750201", "FOYER DE DOMAGNE"),
        ("6750210", "FOYER DE LASSIGNIEU"),
        ("6750207", "FOYER LE SOUS BOIS"),
        ("6750204", "FOYER LE VILLARDOIS"),
        ("6750202", "FOYER LES 4 VENTS"),
        ("6750209", "FOYER LES FLORALIES"),
        ("6750211", "FOYER LES PATIOS"),
        ("6750206", "FOYER LES PRES DE BROU"),
        ("6750214", "FOYER LES SOURDIERES"),
        ("6750208", "FOYER LE VAL FLEURI"),
        ("6750300", "CHAMP D'OR"),
        ("6750102", "IME GEORGES LOISEAU"),
        ("6750105", "IME L'ARMAILLOU"),
        ("6750101", "IME LE PRELION"),
        ("6750103", "IME LES SAPINS"),
        ("6750402", "EA DE BROU"),
        ("6750104", "IME SERVICE LES MUSCARIS"),
        ("6750401", "EA MAISONNETTE"),
        ("6750403", "EA MAISON DES PAYS DE L'AIN"),
        ("6750505", "MAS BELLEVUE"),
        ("6750502", "MAS LES MONTAINES"),
        ("6750501", "MAS MONTPLAISANT"),
        ("6750205", "SAVS LE PASSAGE BG EN B"),
        ("6750001", "ADAPEI DE L'AIN SIEGE SOCIAL"),
        ("6750007", "PCPE"),
        ("6750004", "POLE GEST BOURG EN BRESSE"),
        ("6750005", "POLE DE GESTION OYONNAX"),
        ("6750006", "POLE DE GESTION BELLEY"),
        ("6750003", "POLE GEST FONC TRANSVERSES"),

        # --- SAJ ---
        ("675020902", "SAJ FOYER LES FLORALIES"),
        ("675020102", "SAJ DE DOMAGNE"),
        ("675021402", "SAJ FOYER LES SOURDIERES"),
        ("675020702", "SAJ FOYER SOUS BOIS"),
        ("675021202", "SAJ FOYER DE TREFFORT"),
        ("675020402", "SAJ FOYER LE VILLARDOIS"),
        ("675021002", "SAJ FOYER DE LASSIGNIEU"),

        # --- SAVS ---
        ("675020903", "SAVS FOYER LES FLORALIES"),
        ("675021003", "SAVS FOYER DE LASSIGNIEU"),
        ("675020703", "SAVS SOUS-BOIS"),

        # --- SESSAD ---
        ("675010101", "SESSAD LES DOMBES"),
        ("675010501", "SESSAD INTERLUDE"),
        ("675010201", "SESSAD G LOISEAU"),
        ("67510301",  "SESSAD LES SAPINS"),
    ]

    df = pd.DataFrame(data, columns=["SIRH_ID", "SIRH_NOM"])

    mapping = {}

    for sirh_id, nom in data:
        sirh_id = str(sirh_id)

        # mapping exact (codes longs)
        mapping[sirh_id] = nom

        # fallback sur les 3 derniers chiffres
        if len(sirh_id) >= 3:
            mapping[sirh_id[-3:]] = nom

    return mapping

STRUCTURE_MAP = get_structure_mapping()
# =========================
#   PIPELINE PRINCIPAL
# =========================
def traitement_partie1(fichier_initial: str) -> str:
    """
    1) Extrait remplaçants -> CSV
    2) Copie filtrée (retire dates/headers parasites)
    3) Insère lignes Nom/Prénom
    4) Remplit Nom/Prénom + remplaçants (avec année auto)
    5) Mise en forme & reconstruction finale
    """
    # sorties
    fichier_csv = "fichier_intermediaire.csv"
    fichier_nettoye = "planning_filtre.xlsx"
    fichier_nom_prenom = "planning_avec_nom_prenom.xlsx"
    fichier_final = "planning_final_complet.xlsx"

    # ---------- 1) Extraction remplaçants (col.A) ----------
    df_rempla = extract_remplacants_from_colA(fichier_initial)
    if df_rempla.empty:
        df_rempla = pd.DataFrame(columns=["date", "groupe", "nom", "prenom"])
    df_rempla.to_csv(fichier_csv, index=False, encoding="utf-8")

    # ---------- 2) Copie filtrée ----------
    wb = load_workbook(fichier_initial)
    ws = wb.active
    wb_nouveau = Workbook()
    ws_nouveau = wb_nouveau.active
    ligne_nouvelle = 1
    headers_to_skip = {"nom/", "prénom", "prenom"}

    for row in ws.iter_rows():
        v0 = row[0].value
        if isinstance(v0, str):
            v = v0.strip().lower()
            if re.match(r"\d{2}/\d{2}/\d{4}", v) or v in headers_to_skip:
                continue
        for col_index, cell in enumerate(row, start=1):
            nc = ws_nouveau.cell(row=ligne_nouvelle, column=col_index, value=cell.value)
            if cell.has_style:
                nc.font = copy(cell.font)
                nc.border = copy(cell.border)
                nc.fill = copy(cell.fill)
                nc.number_format = copy(cell.number_format)
                nc.protection = copy(cell.protection)
                nc.alignment = copy(cell.alignment)
        ligne_nouvelle += 1
    wb_nouveau.save(fichier_nettoye)

    # ---------- 3) Insertion lignes Nom/Prénom sous chaque 'Act. jour' ----------
    wb = load_workbook(fichier_nettoye)
    ws = wb.active
    nettoyer_nom_ligne4(ws, col_debut=5)

    lignes_act_jour = [
        r for r in range(1, ws.max_row + 1)
        if isinstance(ws.cell(r, 4).value, str)
        and ws.cell(r, 4).value.strip().lower() == "act. jour"
    ]

    decalage = 0
    for ligne in lignes_act_jour:
        i = ligne + 1 + decalage
        ws.insert_rows(i, amount=2)
        ws.cell(row=i,   column=4, value="Nom").font    = Font(name="Segoe UI", size=14)
        ws.cell(row=i+1, column=4, value="Prénom").font = Font(name="Segoe UI", size=14)
        decalage += 2

    # nettoyage ciblé
    nettoyer_prenom_dans_ligne_nom(ws, start_col=5)
    wb.save(fichier_nom_prenom)

    # ---------- 4) Remplissage (année auto) ----------
    ws = load_workbook(fichier_nom_prenom).active
    year_ref = guess_year_from_column_a(load_workbook(fichier_initial).active, default_year="2026")

    # colonnes jours (E→…)
    DAY_COLS = detect_day_columns_p1(ws, start_col=5)
    last_day_col = max(DAY_COLS) if DAY_COLS else 34
    colonnes = DAY_COLS if DAY_COLS else range(5, 35)

    # conversion entête → date dd/mm/YYYY (avec year_ref)
    def header_to_date(cell_val):
        return parse_header_to_date(cell_val, year=year_ref)

    # pré-normalise groupe CSV
    if not df_rempla.empty:
        df_rempla["_g_norm"] = df_rempla["groupe"].map(norm_group)
        df_rempla["_date_norm"] = df_rempla["date"].map(lambda x: str(x).strip())

    for row in range(2, ws.max_row):
        identite = ws.cell(row=row, column=1).value
        if not isinstance(identite, str) or not identite.strip():
            continue

        r_hor, r_nom, r_pre = row, row + 3, row + 4
        #is_rempla = noacc_lower(identite).startswith("remplac")
        # 🔁 NOUVELLE LOGIQUE REMPLAÇANT : CATEGORIE == 2
        val_categorie = ws.cell(row=row, column=3).value  # colonne C = CATEGORIE
        try:
            is_rempla = int(val_categorie) == 2
        except (TypeError, ValueError):
            is_rempla = False
        groupe_xlsx = norm_group(identite)
        
        # ⚠️ DOIT ÊTRE TOUJOURS DÉFINI
        if not is_rempla:
            # personnes : "NOM\nPrénom" ou "NOM Prénom"
            if "\n" in identite:
                nom_line, prenom_line = identite.split("\n", 1)
                nom = nom_line.strip()
                prenom = prenom_line.strip()
            else:
                parts = identite.strip().split()
                nom = parts[0] if parts else ""
                prenom = " ".join(parts[1:]) if len(parts) > 1 else ""
            # réécrit la cellule A (NOM sur 1ere ligne / Prénom sur 2e)
            ws.cell(r_hor, 1, f"{nom}\n{prenom}").alignment = Alignment(wrap_text=True)
            # remplit lignes Nom/Prénom
            continue
            #for col in colonnes:
            #    for rr, val in zip([r_nom, r_pre], [nom, prenom]):
            #        c = ws.cell(rr, col, val)
            #        c.font = Font(name="Segoe UI", size=8)
            #        c.alignment = Alignment(horizontal="center")
        else:
            # bloc remplaçant : on matche par groupe + date
            for col in colonnes:
                d = header_to_date(ws.cell(1, col).value)
                if not d or df_rempla.empty:
                    continue
                subset = df_rempla[(df_rempla["_g_norm"] == groupe_xlsx) & (df_rempla["_date_norm"] == d)]
                if not subset.empty:
                    nom_csv = strip_placeholders(str(subset.iloc[0]["nom"])).strip()
                    prenom_csv = strip_placeholders(str(subset.iloc[0]["prenom"])).strip()
                    for rr, val in zip([r_nom, r_pre], [nom_csv, prenom_csv]):
                        c = ws.cell(rr, col, val)
                        c.font = Font(name="Segoe UI", size=8)
                        c.alignment = Alignment(horizontal="center")

    # ---------- 5) Mise en forme + reconstruction bornée ----------
    # ajustement des horaires (sauts de ligne)
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 4).value == "Hor.":
            max_ligne = 40
            for col in colonnes:
                cell = ws.cell(row, col)
                if isinstance(cell.value, str):
                    txt = re.sub(r"\s*-\s*", " -\n", cell.value.strip())
                    txt = txt.replace("/", "/\n")
                    cell.value = txt
                    cell.alignment = Alignment(wrap_text=True, horizontal="center")
                    if "/\n" in txt or txt.count("\n") > 1:
                        max_ligne = 80
            ws.row_dimensions[row].height = max_ligne

    # reconstruction
    wb_new = Workbook()
    ws_new = wb_new.active
    r_new = 1
    for row in ws.iter_rows(min_col=1, max_col=last_day_col):
        if all((cell.value in [None, ""]) for cell in row):
            continue
        for col_index, cell in enumerate(row, start=1):
            nc = ws_new.cell(row=r_new, column=col_index, value=cell.value)
            if cell.has_style:
                nc.font = copy(cell.font)
                nc.border = copy(cell.border)
                nc.fill = copy(cell.fill)
                nc.number_format = copy(cell.number_format)
                nc.protection = copy(cell.protection)
                nc.alignment = copy(cell.alignment)
        r_new += 1

    # fusions par bloc (A/B/C)
    for row in range(1, ws_new.max_row - 3):
        if ws_new.cell(row=row, column=4).value == "Hor.":
            for col in [1, 2, 3]:
                ws_new.merge_cells(start_row=row, end_row=row + 4, start_column=col, end_column=col)
                ws_new.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws_new.column_dimensions["A"].width = 50
    nettoyer_nom_ligne4(ws_new, col_debut=5)
    nettoyer_prenom_dans_ligne_nom(ws_new, start_col=5)
    ws_new.column_dimensions["B"].width = 20   # 🔥 ETAB_01 / Groupe
    wb_new.save(fichier_final)
    return fichier_final



def guess_code_from_B2_B6(ws):
    """
    Cherche le code structure (3 chiffres) dans B2..B6.
    Accepte '501 AZUR', '501\nAZUR', 501 (numérique), etc.
    """
    for r in range(2, 7):  # B2 -> B6
        v = ws.cell(row=r, column=2).value  # colonne B = 2
        if v is None:
            continue
        # si Excel a un nombre (501, 501.0)
        if isinstance(v, (int, float)):
            n = int(v)
            if 0 <= n <= 999:
                return f"{n:03d}"
        # sinon chaîne
        s = str(v).replace("\n", " ").strip()
        m = re.search(r"\b(\d{3})\b", s)
        if m:
            return m.group(1)
    return None  # pas trouvé



STRUCTURE_MAP = get_structure_mapping()

# === Helpers entêtes & jours ===
def _norm_text(s: str) -> str:
    s = s.strip().replace("\n", " ").replace(".", " ")
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).lower()

_MOIS_MAP = {
    "jan": "01", "fev": "02", "fév": "02", "mar": "03", "mars": "03",
    "avr": "04", "mai": "05", "juin": "06", "jun": "06",
    "jui": "07", "juil": "07",
    "aou": "08", "aout": "08", "août": "08",
    "sep": "09", "sept": "09",
    "oct": "10", "nov": "11",
    "dec": "12", "déc": "12"
}


def parse_header_to_date(header_val, year="2026"):
    if not isinstance(header_val, str) or not header_val.strip():
        return None
    s = _norm_text(header_val)
    m = re.search(r"(\d{1,2})\s*([a-z]{3,5})", s)
    if not m:
        return None
    j = int(m.group(1))
    mois_tok = m.group(2)[:4]
    if mois_tok.startswith("jui"): mois_tok = "jui"
    if mois_tok.startswith("aou"): mois_tok = "aou"
    if mois_tok.startswith("dec"): mois_tok = "dec"
    mois = _MOIS_MAP.get(mois_tok)
    if not mois:
        return None
    return f"{j:02d}/{mois}/{year}"

def detect_layout(ws):
    # Essaie OFFSET=1 (D=4) puis OFFSET=0 (C=3) pour la colonne "Hor."
    for col_labels in (4, 3):
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=col_labels).value == "Hor.":
                return col_labels
    return 4


def detect_day_columns(ws, col_labels):
    """
    Détecte les colonnes jours à partir de la ligne 1, juste après COL_LABELS.
    Tolère des colonnes vides ou non-date (ex: cellules fusionnées) et s'arrête
    seulement après quelques vides consécutifs.
    """
    day_cols = []
    consecutive_blanks = 0
    for col in range(col_labels + 1, ws.max_column + 1):
        raw = ws.cell(row=1, column=col).value

        # Tolérer des cellules vides/parties de fusion
        if raw in (None, ""):
            consecutive_blanks += 1
            if consecutive_blanks >= 3:   # 3 vides d'affilée ⇒ on considère que c'est fini
                break
            continue

        consecutive_blanks = 0  # on a vu quelque chose

        # Si c'est une date valide → garder la colonne ; sinon ignorer et continuer
        if parse_header_to_date(str(raw)) is not None:
            day_cols.append(col)
        # else: on continue sans breaker (peut être un libellé parasite)

    return day_cols

# === Normalisation robuste des placeholders "Nom/Prénom" ===
def normalize_header(s: str) -> str:
    s = s.strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    s = s.replace(" ", "")
    return s

def clean_cell_header_like(v):
    """Retourne '' si la cellule est un placeholder ('Nom', 'Prénom', 'Nom/Prénom' avec variantes)."""
    if v is None:
        return ""
    s = str(v)
    s_norm = normalize_header(s)
    if s_norm in {"nom", "prenom", "nom/prenom", "nom/prénom", "nom/prenom/", "nom/"}:
        return ""
    return s.strip()
from openpyxl import load_workbook

def read_equipe_label(xlsx_path: str) -> str:
    """
    Lit la première valeur non vide trouvée en colonne B (lignes 2..10) sur le 1er onglet.
    Ex: '501 AZUR' (ou '501\\nAZUR'). On normalise juste les espaces.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws0 = wb.worksheets[0]  # premier onglet (“sheet”)
    for r in range(2, 11):
        v = ws0.cell(row=r, column=2).value  # colonne B
        if v:
            s = str(v).strip().replace("\n", " ")
            s = " ".join(s.split())  # compacter espaces
            # 👉 Si tu ne veux que le libellé sans les 3 chiffres, décommente:
            # m = re.search(r"\b\d{3}\s*(.*)", s)
            # if m: s = m.group(1).strip()
            return s
    return ""

def read_structure_from_lecture(xlsx_path: str) -> str:
    """
    Lit le titre (structure) depuis l’onglet 'lecture', cellule C3.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    if "lecture" not in wb.sheetnames:
        return ""

    ws = wb["lecture"]
    val = ws.cell(row=3, column=3).value  # C3
    return str(val).strip() if val else ""

# === PARTIE 2 : Générer les onglets "lecture" et "interimaire" ===

def traitement_partie2(fichier_source: str) -> str:
    wb = load_workbook(fichier_source)

    # 1) Supprimer toutes les anciennes tables (sinon Excel casse les formules)
    for ws in wb.worksheets:
        for tbl in list(ws._tables):
            ws.remove_table(tbl.name)

    # 2) Supprimer anciens onglets lecture / interimaire
    for name in ["lecture", "interimaire"]:
        if name in wb.sheetnames:
            wb.remove(wb[name])

    ws_source = wb.active

    # === Détection structurée ===
    COL_LABELS = detect_layout(ws_source)
    DAY_COLS   = detect_day_columns(ws_source, COL_LABELS)

    dates_colonnes = {
        col: parse_header_to_date(ws_source.cell(row=1, column=col).value, "2026")
        for col in DAY_COLS
        if parse_header_to_date(ws_source.cell(row=1, column=col).value, "2026") is not None
    }

    max_row = ws_source.max_row

    lignes_hor = [
        r for r in range(1, max_row + 1)
        if ws_source.cell(row=r, column=COL_LABELS).value == "Hor."
    ]

    lignes_donnees = []

    for ligne_hor in lignes_hor:
        ligne_lieu   = min(ligne_hor + 1, max_row)
        ligne_act    = min(ligne_hor + 2, max_row)
        ligne_nom    = min(ligne_hor + 3, max_row)
        ligne_prenom = min(ligne_hor + 4, max_row)

        valeur_nom_colA = ws_source.cell(row=ligne_hor, column=1).value or ""

        for col in DAY_COLS:
            val_act  = ws_source.cell(row=ligne_act,  column=col).value
            val_hor  = ws_source.cell(row=ligne_hor,  column=col).value
            val_lieu = ws_source.cell(row=ligne_lieu, column=col).value

            #if not isinstance(val_act, str):
            #    continue
#
            #val_norm = re.sub(r"\s+", "", val_act.upper())
            #if not re.match(r"^\d{3}(G\d{1,2}|[A-Z]\d{0,2})$", val_norm):
            #    continue

            # --------------------------------------------------
            # GROUPE : priorité à Act. jour, sinon fallback col B
            # --------------------------------------------------
            groupe_val = None

            if isinstance(val_act, str) and val_act.strip():
                candidate = re.sub(r"\s+", "", val_act.upper())
                if re.match(r"^\d{3}(G\d{1,2}|[A-Z]\d{0,2})$", candidate):
                    groupe_val = candidate

            # fallback : colonne B (ETAB_01)
            if not groupe_val:
                val_b = ws_source.cell(row=ligne_hor, column=2).value  # colonne B
                if val_b:
                    groupe_val = str(val_b).strip()


            if isinstance(val_hor, str):
                hor_clean = val_hor.replace("\n", "").replace(" ", "")
                if re.match(r"^0{1,2}:0{2}-0{1,2}:0{2}$", hor_clean):
                    continue

            date_cell = dates_colonnes.get(col, "")

            brut_nom    = ws_source.cell(row=ligne_nom,    column=col).value
            brut_prenom = ws_source.cell(row=ligne_prenom, column=col).value

            nom_clean    = clean_cell_header_like(brut_nom)
            prenom_clean = clean_cell_header_like(brut_prenom)

            nom_concat = " ".join(x for x in [nom_clean, prenom_clean] if x).strip()
            if not nom_concat:
                nom_concat = str(valeur_nom_colA).replace("\n", " ").strip()

            lignes_donnees.append([
                date_cell,
                groupe_val,
                val_hor,
                "",
                "",
                nom_concat,
                val_lieu
            ])

    # === TRI ===
    lignes_donnees = sorted(
        lignes_donnees,
        key=lambda x: (pd.to_datetime(x[0], dayfirst=True, errors="coerce"), x[1])
    )

    entetes = [
        "Date", "Groupe", "Horaire",
        "Motif", "NOM de la personne remplacée",
        "Nom", "Agence"
    ]

    #code_titre = guess_code_from_B2_B6(ws_source)
    #if not code_titre and lignes_donnees:
    #    g = lignes_donnees[0][1]
    #    if g[:3].isdigit():
    #        code_titre = g[:3]
    #if not code_titre:
    #    code_titre = "502"


    # 🔹 Titre basé sur ETAB_01 (colonne B)
    code_titre = None

    for r in range(2, 7):  # B2 → B6
        v = ws_source.cell(row=r, column=2).value
        if v:
            code_titre = str(v).strip()
            break

    # fallback ultra sécurisé
    if not code_titre and lignes_donnees:
        g = lignes_donnees[0][1]
        code_titre = str(g)

    # dernier filet
    if not code_titre:
        code_titre = "UNKNOWN"


    # ---------------------------------------------------------------------
    #  📌 *** NOUVEAU : DataFrame avec row_source = ligne exacte dans lecture ***
    # ---------------------------------------------------------------------
    df = pd.DataFrame(lignes_donnees, columns=entetes)
    df["row_source"] = df.index + 5   # 5 = première ligne d’écriture dans lecture

    # === FILTRAGE INTERIMAIRE ===
    df_interim = df[df["Agence"] == "A POURVOIR"].copy()
    df_interim["Date"] = pd.to_datetime(df_interim["Date"], dayfirst=True, errors="coerce")
    df_interim = df_interim.sort_values(by=["Date", "Groupe"])
    df_interim["Date"] = df_interim["Date"].dt.strftime("%d/%m/%Y")

    # Colonnes à envoyer à la fonction créer_onglet
    lignes_interim = df_interim[
        ["Date", "Groupe", "Horaire", "Motif",
         "NOM de la personne remplacée", "Nom", "Agence", "row_source"]
    ].values.tolist()

        # --------------------------------------------------
# Détection du suffixe de groupe (G1, G2, ...)
# --------------------------------------------------
    suffix_groupe = ""

    for row in lignes_donnees:
        grp = str(row[1])  # colonne "Groupe"
        m = re.search(r"G\d+", grp)
        if m:
            suffix_groupe = " " + m.group(0)
            break  # on prend le premier groupe trouvé

    # ---------------------------------------------------------------------
    #  FONCTION CRÉATION D’ONGLET AVEC MIRROIR FONCTIONNEL
    # ---------------------------------------------------------------------

    def creer_onglet(nom_onglet, lignes, code_structure):
        code_structure = re.sub(r"\D", "", str(code_structure))
        nom_structure = STRUCTURE_MAP.get(code_structure, f"Structure {code_structure}")

        titre_complet = f"{nom_structure}{suffix_groupe}".upper()

        ws = wb.create_sheet(nom_onglet)

        # =========================
        # Titre
        # =========================
        ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=9)
        t = ws.cell(row=3, column=3, value=titre_complet)
        t.font = Font(name="Aptos Narrow", size=11, bold=True)
        t.alignment = Alignment(horizontal="center")
        t.fill = PatternFill("solid", start_color="FBE2D5")

        # =========================
        # En-têtes
        # =========================
        for idx, val in enumerate(entetes, start=3):
            c = ws.cell(row=4, column=idx, value=val)
            c.font = Font(name="Aptos Narrow", size=11, bold=True)
            c.alignment = Alignment(horizontal="center")

        # =========================
        # Écriture des lignes
        # =========================
        ligne = 5
        for row in lignes:
            *data, row_src = row
            for col_idx, val in enumerate(data, start=3):
                cell = ws.cell(row=ligne, column=col_idx)

                if nom_onglet == "interimaire" and col_idx in (6, 7):
                    col_letter = "F" if col_idx == 6 else "G"
                    cell.value = f"=lecture!{col_letter}{row_src}"
                else:
                    cell.value = val

                cell.font = Font(name="Aptos Narrow", size=11)
                cell.alignment = Alignment(horizontal="center")

            ligne += 1

        # =========================
        # Calcul last_row (OBLIGATOIRE AVANT USAGE)
        # =========================
        last_row = ligne - 1


        # =========================
        # Table Excel (APRÈS validation)
        # =========================
        if last_row >= 5:
            ref = f"C4:I{last_row}"
            table = Table(displayName=f"Table_{nom_onglet}", ref=ref)
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2")
            ws.add_table(table)

        # =========================
        # Menu déroulant "Motif" (colonne F)
        # =========================
        if nom_onglet == "lecture" and last_row >= 5:
            motifs_str = ",".join(MOTIFS_PREDEFINIS)

            dv = DataValidation(
                type="list",
                formula1=f'"{motifs_str}"',
                allow_blank=True,
                showDropDown=False
            )

            ws.add_data_validation(dv)
            dv.add(f"F5:F{last_row}")
        # DEBUG : écrire un marqueur visible
        #ws["F5"].value = "DEBUG_MOTIF"
        #ws["G5"].value = "DEBUG_REMPLACE"
        # =========================
        # Largeurs colonnes
        # =========================
        ws.column_dimensions["B"].width = 20

        for col_idx in range(3, 10):
            ws.column_dimensions[chr(64 + col_idx)].width = 20

        # 🎯 Élargir la colonne F uniquement pour lecture & interimaire
        if nom_onglet in ("lecture", "interimaire"):
            ws.column_dimensions["B"].width = 10
            ws.column_dimensions["F"].width = 32
            ws.column_dimensions["E"].width = 25
            ws.column_dimensions["G"].width = 28
            ws.column_dimensions["H"].width = 25


    # === Création lecture ===
    creer_onglet("lecture", df.values.tolist(), code_titre)

    # === Création interimaire avec miroir fiable ===
    creer_onglet("interimaire", lignes_interim, code_titre)

    fichier_sortie = "planning_final_complet.xlsx"
    wb.save(fichier_sortie)
    return fichier_sortie

# === PARTIE 3 : Générer le fichier Badakan à partir de l'onglet "interimaire" ===
def adapter_badakan_version_auto(df_interim: pd.DataFrame,
                                 structure_label: str) -> pd.DataFrame:

    jours_fr = {
        "Monday": "Lundi", "Tuesday": "Mardi", "Wednesday": "Mercredi",
        "Thursday": "Jeudi", "Friday": "Vendredi", "Saturday": "Samedi", "Sunday": "Dimanche"
    }

    def fix_time(t: str) -> str:
        t = t.strip()
        if ":" in t:
            m = re.fullmatch(r"(\d{1,2}):(\d{1,2})", t)
            if m:
                return f"{int(m.group(1)):02d}:{int(m.group(2)):02d}"
            return t
        m = re.fullmatch(r"(\d{1,2})(\d{2})", t)
        if m:
            return f"{int(m.group(1)):02d}:{m.group(2)}"
        return t

    def calculer_infos_horaires(horaire_str: str):
        if not isinstance(horaire_str, str) or not horaire_str.strip():
            return "", "", "00:00", 0.0

        hs = horaire_str.replace("\n", "").replace(" ", "")
        segments = [s for s in hs.split("/") if s]
        plages = []
        for seg in segments:
            if "-" in seg:
                d, f = seg.split("-", 1)
                plages.append((fix_time(d), fix_time(f)))

        if not plages:
            return "", "", "00:00", 0.0

        heure_debut = plages[0][0]
        heure_fin   = plages[-1][1]

        def to_dt(hhmm: str) -> datetime:
            return datetime.strptime(hhmm, "%H:%M")

        total_travail_min = 0
        for d, f in plages:
            total_travail_min += int((to_dt(f) - to_dt(d)).seconds / 60)

        amplitude_min = int((to_dt(heure_fin) - to_dt(heure_debut)).seconds / 60)
        coupure_min = max(amplitude_min - total_travail_min, 0)
        t_travail_dec = round(total_travail_min / 60, 2)

        def min_to_hhmm(m: int) -> str:
            h, mn = divmod(m, 60)
            return f"{h:02d}:{mn:02d}"

        return heure_debut, heure_fin, min_to_hhmm(coupure_min), t_travail_dec

    df_interim["Date_dt"] = pd.to_datetime(df_interim["Date"], dayfirst=True, errors="coerce")
    df_interim = df_interim.sort_values(by=["Date_dt", "Groupe"]).reset_index(drop=True)

    # Prénoms dynamiques (Interimaire_1, Interimaire_2 par date)
    df_interim["Prénom"] = (df_interim.groupby("Date_dt").cumcount() + 1).apply(lambda x: f"Interimaire_{x}")

    results = []
    for _, row in df_interim.iterrows():
        groupe_val = str(row["Groupe"]).strip()
        #code_struct = groupe_val[:3] if len(groupe_val) >= 4 else "000"

        # 👉 Nouveau: on privilégie l’étiquette lue en B (onglet 1)
        unite_val = groupe_val 
        structure = structure_label

        try:
            date_obj = pd.to_datetime(row["Date"], dayfirst=True)
            jour_fr = jours_fr[date_obj.strftime("%A")]
            date_str = f"{jour_fr} {date_obj.strftime('%d/%m/%Y')}"
        except Exception:
            date_str = ""

        horaire = str(row["Horaire"])
        heure_debut, heure_fin, coupure, t_travail = calculer_infos_horaires(horaire)

        results.append({
            "Nom": "Interimaire",
            "Prénom": row["Prénom"],
            "Poste": "Accompagnant éducatif et soc",
            "Stucture(s)": structure,
            "Date": date_str,
            "Heure de début de travail": heure_debut,
            "Temps de coupure": coupure,
            "Heure de fin de travail": heure_fin,
            "Temps travaillé": str(t_travail).replace(".", ","),
            "Personne remplacée": row.get("Personne remplacée", ""),
            "Motif": row.get("Motif", ""),
            "Info complémentaire sur le motif": "",
            "Unite(s)": unite_val,
            "Précisez si coefficient EXTERNAT": "",
            "Commentaires": ""
        })

    return pd.DataFrame(results)

def traitement_partie3(fichier_interimaire: str) -> str:
    structure_label = read_structure_from_lecture(fichier_interimaire)
    # Lit l’onglet "interimaire" produit par traitement_partie2

    df_interim = pd.read_excel(
        fichier_interimaire,
        sheet_name="interimaire",
        skiprows=3,
        usecols="C:I"
    )
    df_interim.columns = [
        "Date", "Groupe", "Horaire", "Motif",
        "Personne remplacée", "Nom", "Agence"
    ]

    # Passer l’étiquette à l’adaptateur
    df_badakan = adapter_badakan_version_auto(
        df_interim,
        structure_label=structure_label
    )
    fichier_badakan = "badakan.csv"
    df_badakan.to_csv(fichier_badakan, sep=';', index=False,  header=True,   encoding='utf-8-sig')
    return fichier_badakan


def traitement_pipeline_complet(fichier_brut: str) -> str:
    """
    Pipeline complet :
    1) Traitement partie 1 (nettoyage + planning stylisé)
    2) Traitement partie 2 (lecture + interimaire)
    """
    # Étape 1
    fichier_p1 = traitement_partie1(fichier_brut)

    # Étape 2
    fichier_p2 = traitement_partie2(fichier_p1)

    return fichier_p2

# =========================
#   UI
# =========================
st.header("1️⃣ Planning, Lecture & Intérimaire")

uploaded_file_full = st.file_uploader(
    " ",
    type=["xlsx"],
    key="upload_full"
)

if uploaded_file_full and st.button("Générer planning + lecture + intérimaire"):
    raw_path = save_uploaded_file(uploaded_file_full, "raw")

    with st.spinner("Pipeline complet en cours…"):
        fichier_final = traitement_pipeline_complet(raw_path)

    st.success("✅ Pipeline complet terminé")

    with open(fichier_final, "rb") as f:
        st.download_button(
            "📥 Télécharger le fichier final",
            data=f,
            file_name=os.path.basename(fichier_final)
        )




st.header("2️⃣ Génération du fichier Badakan")

uploaded_file_3 = st.file_uploader(
    " ",
    type=["xlsx"],
    key="upload3"
)

if st.button("Générer le fichier Badakan"):
    if uploaded_file_3:
        source_p3 = save_uploaded_file(uploaded_file_3, "interimaire")
    else:
        source_p3 = st.session_state.get("planning_p2")

    if not source_p3:
        st.error("❌ Aucun fichier interimaire disponible")
    else:
        with st.spinner("Génération du fichier Badakan…"):
            badakan = traitement_partie3(source_p3)

        st.success("✅ Fichier Badakan généré")
        st.download_button(
            "📥 Télécharger Badakan.csv",
            data=open(badakan, "rb"),
            file_name="badakan.csv"
        )
